#!/usr/bin/env python3
"""
Улучшенная функция экспорта Excel с графиками и детальной аналитикой
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.drawing.image import Image
from io import BytesIO
import json
from datetime import datetime, timedelta

# Импортируем модели базы данных
try:
    from app import Answer
except ImportError:
    # Если импорт не работает, создаем заглушку
    class Answer:
        @staticmethod
        def query():
            return None

def create_enhanced_excel_report(survey, responses, question_analytics):
    """Создает улучшенный Excel отчет с графиками и аналитикой"""
    
    wb = Workbook()
    
    # Стили
    header_font = Font(bold=True, size=16, color='FFFFFF')
    header_fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    subheader_font = Font(bold=True, size=12, color='2C3E50')
    subheader_fill = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
    
    data_font = Font(size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    number_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # ========== ЛИСТ 1: ИСПОЛНИТЕЛЬНОЕ РЕЗЮМЕ ==========
    ws_summary = wb.active
    ws_summary.title = 'Исполнительное резюме'
    
    # Заголовок
    ws_summary.merge_cells('A1:H1')
    ws_summary['A1'] = f'ИСПОЛНИТЕЛЬНОЕ РЕЗЮМЕ: {survey.title}'
    ws_summary['A1'].font = header_font
    ws_summary['A1'].fill = header_fill
    ws_summary['A1'].alignment = header_alignment
    ws_summary.row_dimensions[1].height = 35
    
    # Ключевые метрики
    row = 3
    metrics = [
        ('Общее количество ответов', len(responses)),
        ('Количество вопросов', len(survey.questions)),
        ('Дата создания', survey.created_at.strftime('%d.%m.%Y')),
        ('Статус', 'Активен' if survey.is_active else 'Неактивен'),
        ('Тип опроса', 'Анонимный' if survey.is_anonymous else 'Авторизованный'),
        ('Среднее время прохождения', calculate_avg_completion_time(responses)),
        ('Процент завершения', calculate_completion_rate(survey, responses)),
        ('Уникальные IP адреса', len(set(r.ip_address for r in responses if r.ip_address)))
    ]
    
    for i, (label, value) in enumerate(metrics):
        col = (i % 4) * 2 + 1
        current_row = row + (i // 4) * 2
        
        ws_summary[f'{chr(64+col)}{current_row}'] = label
        ws_summary[f'{chr(64+col)}{current_row}'].font = subheader_font
        ws_summary[f'{chr(64+col)}{current_row}'].fill = subheader_fill
        ws_summary[f'{chr(64+col)}{current_row}'].border = border
        
        ws_summary[f'{chr(65+col)}{current_row}'] = value
        ws_summary[f'{chr(65+col)}{current_row}'].font = data_font
        ws_summary[f'{chr(65+col)}{current_row}'].alignment = number_alignment
        ws_summary[f'{chr(65+col)}{current_row}'].border = border
    
    # Ключевые инсайты
    row += 8
    ws_summary.merge_cells(f'A{row}:H{row}')
    ws_summary[f'A{row}'] = 'КЛЮЧЕВЫЕ ИНСАЙТЫ'
    ws_summary[f'A{row}'].font = subheader_font
    ws_summary[f'A{row}'].fill = subheader_fill
    ws_summary[f'A{row}'].border = border
    row += 1
    
    insights = generate_key_insights(survey, responses, question_analytics)
    for insight in insights:
        ws_summary.merge_cells(f'A{row}:H{row}')
        ws_summary[f'A{row}'] = f"• {insight}"
        ws_summary[f'A{row}'].font = data_font
        ws_summary[f'A{row}'].border = border
        row += 1
    
    # Рекомендации
    row += 1
    ws_summary.merge_cells(f'A{row}:H{row}')
    ws_summary[f'A{row}'] = 'РЕКОМЕНДАЦИИ'
    ws_summary[f'A{row}'].font = subheader_font
    ws_summary[f'A{row}'].fill = subheader_fill
    ws_summary[f'A{row}'].border = border
    row += 1
    
    recommendations = generate_recommendations(survey, responses, question_analytics)
    for rec in recommendations:
        ws_summary.merge_cells(f'A{row}:H{row}')
        ws_summary[f'A{row}'] = f"• {rec}"
        ws_summary[f'A{row}'].font = data_font
        ws_summary[f'A{row}'].border = border
        row += 1
    
    # Устанавливаем ширину колонок
    for col in range(1, 9):
        ws_summary.column_dimensions[chr(64+col)].width = 15
    
    # ========== ЛИСТ 2: ДЕТАЛЬНАЯ АНАЛИТИКА ==========
    ws_analytics = wb.create_sheet('Детальная аналитика')
    
    # Заголовок
    ws_analytics.merge_cells('A1:J1')
    ws_analytics['A1'] = 'ДЕТАЛЬНАЯ АНАЛИТИКА ПО ВОПРОСАМ'
    ws_analytics['A1'].font = header_font
    ws_analytics['A1'].fill = header_fill
    ws_analytics['A1'].alignment = header_alignment
    ws_analytics.row_dimensions[1].height = 35
    
    # Заголовки таблицы
    headers = ['Вопрос', 'Тип', 'Ответов', 'Процент ответов', 'Популярный ответ', 
               'Процент популярного', 'Инсайты', 'Рекомендации', 'Средняя длина', 'Вариативность']
    
    for i, header in enumerate(headers, 1):
        cell = ws_analytics.cell(row=3, column=i, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
        cell.alignment = header_alignment
    
    # Данные по вопросам
    row = 4
    for question in survey.questions:
        analytics = question_analytics.get(question.id, {})
        data = analytics.get('data', {})
        insights = analytics.get('insights', [])
        recommendations = analytics.get('recommendations', [])
        
        # Основная информация
        ws_analytics[f'A{row}'] = question.text[:50] + ('...' if len(question.text) > 50 else '')
        ws_analytics[f'A{row}'].font = data_font
        ws_analytics[f'A{row}'].border = border
        
        ws_analytics[f'B{row}'] = get_question_type_name(question.type)
        ws_analytics[f'B{row}'].font = data_font
        ws_analytics[f'B{row}'].border = border
        
        ws_analytics[f'C{row}'] = analytics.get('total_answers', 0)
        ws_analytics[f'C{row}'].font = data_font
        ws_analytics[f'C{row}'].alignment = number_alignment
        ws_analytics[f'C{row}'].border = border
        
        ws_analytics[f'D{row}'] = f"{analytics.get('response_rate', 0):.1f}%"
        ws_analytics[f'D{row}'].font = data_font
        ws_analytics[f'D{row}'].alignment = number_alignment
        ws_analytics[f'D{row}'].border = border
        
        # Популярный ответ
        if question.type in ['single_choice', 'multiple_choice', 'dropdown']:
            most_popular = data.get('most_popular', ('Нет данных', 0))
            ws_analytics[f'E{row}'] = most_popular[0]
            ws_analytics[f'F{row}'] = f"{most_popular[1]} ({most_popular[1]/len(responses)*100:.1f}%)" if len(responses) > 0 else "0%"
        elif question.type in ['rating', 'scale']:
            avg_rating = data.get('avg', 0)
            ws_analytics[f'E{row}'] = f"Средняя оценка: {avg_rating}"
            ws_analytics[f'F{row}'] = f"±{data.get('std_deviation', 0):.1f}"
        else:
            ws_analytics[f'E{row}'] = "Текстовый ответ"
            ws_analytics[f'F{row}'] = f"Средняя длина: {data.get('avg_length', 0):.0f} симв."
        
        ws_analytics[f'E{row}'].font = data_font
        ws_analytics[f'E{row}'].border = border
        ws_analytics[f'F{row}'].font = data_font
        ws_analytics[f'F{row}'].border = border
        
        # Инсайты и рекомендации
        ws_analytics[f'G{row}'] = '; '.join(insights[:2]) if insights else 'Нет данных'
        ws_analytics[f'G{row}'].font = data_font
        ws_analytics[f'G{row}'].border = border
        
        ws_analytics[f'H{row}'] = '; '.join(recommendations[:2]) if recommendations else 'Нет данных'
        ws_analytics[f'H{row}'].font = data_font
        ws_analytics[f'H{row}'].border = border
        
        # Дополнительные метрики
        if question.type in ['text', 'text_paragraph']:
            ws_analytics[f'I{row}'] = f"{data.get('avg_length', 0):.0f}"
            ws_analytics[f'J{row}'] = f"{data.get('max_length', 0) - data.get('min_length', 0):.0f}"
        elif question.type in ['rating', 'scale']:
            ws_analytics[f'I{row}'] = f"{data.get('avg', 0):.1f}"
            ws_analytics[f'J{row}'] = f"{data.get('coefficient_of_variation', 0):.1f}%"
        else:
            ws_analytics[f'I{row}'] = "-"
            ws_analytics[f'J{row}'] = "-"
        
        ws_analytics[f'I{row}'].font = data_font
        ws_analytics[f'I{row}'].alignment = number_alignment
        ws_analytics[f'I{row}'].border = border
        
        ws_analytics[f'J{row}'].font = data_font
        ws_analytics[f'J{row}'].alignment = number_alignment
        ws_analytics[f'J{row}'].border = border
        
        row += 1
    
    # Устанавливаем ширину колонок
    column_widths = [30, 15, 10, 12, 20, 15, 25, 25, 12, 12]
    for i, width in enumerate(column_widths, 1):
        ws_analytics.column_dimensions[chr(64+i)].width = width
    
    # ========== ЛИСТ 3: ГРАФИКИ И ВИЗУАЛИЗАЦИЯ ==========
    ws_charts = wb.create_sheet('Графики и визуализация')
    
    # Заголовок
    ws_charts.merge_cells('A1:H1')
    ws_charts['A1'] = 'ГРАФИКИ И ВИЗУАЛИЗАЦИЯ ДАННЫХ'
    ws_charts['A1'].font = header_font
    ws_charts['A1'].fill = header_fill
    ws_charts['A1'].alignment = header_alignment
    ws_charts.row_dimensions[1].height = 35
    
    # Создаем графики для каждого вопроса
    chart_row = 3
    for question in survey.questions:
        analytics = question_analytics.get(question.id, {})
        data = analytics.get('data', {})
        
        if question.type in ['single_choice', 'multiple_choice', 'dropdown']:
            # Столбчатая диаграмма
            create_bar_chart(ws_charts, question, data, chart_row)
            chart_row += 20
        elif question.type in ['rating', 'scale']:
            # Линейная диаграмма распределения
            create_rating_chart(ws_charts, question, data, chart_row)
            chart_row += 20
        elif question.type == 'checkbox':
            # Круговая диаграмма
            create_pie_chart(ws_charts, question, data, chart_row)
            chart_row += 20
    
    # ========== ЛИСТ 4: ОТВЕТЫ ПОЛЬЗОВАТЕЛЕЙ ==========
    ws_responses = wb.create_sheet('Ответы пользователей')
    
    # Заголовок
    ws_responses.merge_cells('A1:Z1')
    ws_responses['A1'] = 'ДЕТАЛЬНЫЕ ОТВЕТЫ ПОЛЬЗОВАТЕЛЕЙ'
    ws_responses['A1'].font = header_font
    ws_responses['A1'].fill = header_fill
    ws_responses['A1'].alignment = header_alignment
    ws_responses.row_dimensions[1].height = 35
    
    # Заголовки
    headers = ['ID ответа', 'Дата', 'Пользователь', 'IP адрес', 'Время прохождения']
    for i, question in enumerate(survey.questions):
        headers.append(f'Q{i+1}: {question.text[:30]}...')
    
    for i, header in enumerate(headers):
        cell = ws_responses.cell(row=3, column=i+1, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = border
        cell.alignment = header_alignment
    
    # Данные ответов
    row = 4
    for response in responses:
        # Основная информация
        ws_responses[f'A{row}'] = response.id
        ws_responses[f'B{row}'] = response.created_at.strftime('%d.%m.%Y %H:%M')
        
        # Определяем имя пользователя
        if survey.is_anonymous:
            user_name = 'Аноним'
        elif survey.require_name:
            user_name = response.respondent_name or 'Не указано'
        elif response.user_id and response.user:
            user_name = response.user.username
        else:
            user_name = 'Неизвестно'
        
        ws_responses[f'C{row}'] = user_name
        ws_responses[f'D{row}'] = response.ip_address or 'Не указан'
        ws_responses[f'E{row}'] = f"{response.completion_time or 0:.1f} мин"
        
        # Ответы на вопросы
        col = 6
        for question in survey.questions:
            # Ищем ответ через relationship
            answer = None
            if hasattr(response, 'answers'):
                for ans in response.answers:
                    if ans.question_id == question.id:
                        answer = ans
                        break
            
            if answer and answer.value:
                answer_text = format_answer_for_excel(answer, question)
                ws_responses.cell(row=row, column=col, value=answer_text)
            else:
                ws_responses.cell(row=row, column=col, value='Нет ответа')
            col += 1
        
        # Стилизация строки
        for col in range(1, len(headers) + 1):
            cell = ws_responses.cell(row=row, column=col)
            cell.font = data_font
            cell.border = border
            if col <= 5:  # Основные колонки
                cell.alignment = number_alignment
        
        row += 1
    
    # Устанавливаем ширину колонок
    ws_responses.column_dimensions['A'].width = 8
    ws_responses.column_dimensions['B'].width = 15
    ws_responses.column_dimensions['C'].width = 15
    ws_responses.column_dimensions['D'].width = 15
    ws_responses.column_dimensions['E'].width = 12
    for col in range(6, len(headers) + 1):
        ws_responses.column_dimensions[chr(64+col)].width = 20
    
    return wb

def calculate_avg_completion_time(responses):
    """Вычисляет среднее время прохождения"""
    times = [r.completion_time for r in responses if r.completion_time]
    return f"{sum(times)/len(times):.1f} мин" if times else "Нет данных"

def calculate_completion_rate(survey, responses):
    """Вычисляет процент завершения опроса"""
    if not responses:
        return "0%"
    
    total_questions = len(survey.questions)
    completed_responses = 0
    
    for response in responses:
        # Подсчитываем количество ответов через relationship
        answered_questions = len(response.answers) if hasattr(response, 'answers') else 0
        if answered_questions >= total_questions * 0.8:  # 80% вопросов отвечено
            completed_responses += 1
    
    return f"{(completed_responses/len(responses)*100):.1f}%"

def generate_key_insights(survey, responses, question_analytics):
    """Генерирует ключевые инсайты"""
    insights = []
    
    # Общие инсайты
    if len(responses) > 50:
        insights.append(f"Высокая вовлеченность: {len(responses)} ответов")
    elif len(responses) < 10:
        insights.append(f"Низкая вовлеченность: только {len(responses)} ответов")
    
    # Инсайты по вопросам
    for question in survey.questions:
        analytics = question_analytics.get(question.id, {})
        insights.extend(analytics.get('insights', [])[:1])  # Берем только первый инсайт
    
    return insights[:5]  # Максимум 5 инсайтов

def generate_recommendations(survey, responses, question_analytics):
    """Генерирует рекомендации"""
    recommendations = []
    
    # Общие рекомендации
    if len(responses) < 20:
        recommendations.append("Рассмотрите продление срока проведения опроса для увеличения количества ответов")
    
    # Рекомендации по вопросам
    for question in survey.questions:
        analytics = question_analytics.get(question.id, {})
        recommendations.extend(analytics.get('recommendations', [])[:1])  # Берем только первую рекомендацию
    
    return recommendations[:5]  # Максимум 5 рекомендаций

def get_question_type_name(question_type):
    """Возвращает читаемое название типа вопроса"""
    type_names = {
        'text': 'Текст (строка)',
        'text_paragraph': 'Текст (абзац)',
        'single_choice': 'Один из списка',
        'multiple_choice': 'Несколько из списка',
        'dropdown': 'Раскрывающийся список',
        'scale': 'Шкала',
        'rating': 'Оценка',
        'grid': 'Сетка',
        'checkbox_grid': 'Сетка из флажков',
        'date': 'Дата',
        'time': 'Время'
    }
    return type_names.get(question_type, question_type)

def create_bar_chart(worksheet, question, data, start_row):
    """Создает столбчатую диаграмму"""
    if not data.get('counts'):
        return
    
    # Подготавливаем данные
    counts = data['counts']
    sorted_data = sorted(counts.items(), key=lambda x: x[1], reverse=True)[:10]  # Топ-10
    
    # Записываем данные
    worksheet[f'A{start_row}'] = f'Распределение ответов: {question.text[:40]}...'
    worksheet[f'A{start_row}'].font = Font(bold=True, size=12)
    
    for i, (option, count) in enumerate(sorted_data):
        worksheet[f'A{start_row + 2 + i}'] = option[:30]
        worksheet[f'B{start_row + 2 + i}'] = count
    
    # Создаем диаграмму
    chart = BarChart()
    chart.title = f'Ответы на вопрос: {question.text[:30]}...'
    chart.style = 10
    
    data_range = Reference(worksheet, min_col=2, min_row=start_row + 2, 
                          max_row=start_row + 2 + len(sorted_data))
    categories = Reference(worksheet, min_col=1, min_row=start_row + 3, 
                          max_row=start_row + 2 + len(sorted_data))
    
    chart.add_data(data_range)
    chart.set_categories(categories)
    
    worksheet.add_chart(chart, f'D{start_row}')

def create_rating_chart(worksheet, question, data, start_row):
    """Создает диаграмму распределения оценок"""
    if not data.get('distribution'):
        return
    
    distribution = data['distribution']
    
    # Записываем данные
    worksheet[f'A{start_row}'] = f'Распределение оценок: {question.text[:40]}...'
    worksheet[f'A{start_row}'].font = Font(bold=True, size=12)
    
    for rating, count in distribution.items():
        worksheet[f'A{start_row + 2 + int(rating)}'] = f'Оценка {rating}'
        worksheet[f'B{start_row + 2 + int(rating)}'] = count
    
    # Создаем диаграмму
    chart = BarChart()
    chart.title = f'Распределение оценок: {question.text[:30]}...'
    chart.style = 10
    
    data_range = Reference(worksheet, min_col=2, min_row=start_row + 2, 
                          max_row=start_row + 2 + len(distribution))
    categories = Reference(worksheet, min_col=1, min_row=start_row + 3, 
                          max_row=start_row + 2 + len(distribution))
    
    chart.add_data(data_range)
    chart.set_categories(categories)
    
    worksheet.add_chart(chart, f'D{start_row}')

def create_pie_chart(worksheet, question, data, start_row):
    """Создает круговую диаграмму"""
    if not data.get('counts'):
        return
    
    counts = data['counts']
    sorted_data = sorted(counts.items(), key=lambda x: x[1], reverse=True)[:8]  # Топ-8
    
    # Записываем данные
    worksheet[f'A{start_row}'] = f'Распределение выборов: {question.text[:40]}...'
    worksheet[f'A{start_row}'].font = Font(bold=True, size=12)
    
    for i, (option, count) in enumerate(sorted_data):
        worksheet[f'A{start_row + 2 + i}'] = option[:30]
        worksheet[f'B{start_row + 2 + i}'] = count
    
    # Создаем диаграмму
    chart = PieChart()
    chart.title = f'Выборы: {question.text[:30]}...'
    
    data_range = Reference(worksheet, min_col=2, min_row=start_row + 2, 
                          max_row=start_row + 2 + len(sorted_data))
    categories = Reference(worksheet, min_col=1, min_row=start_row + 3, 
                          max_row=start_row + 2 + len(sorted_data))
    
    chart.add_data(data_range)
    chart.set_categories(categories)
    
    worksheet.add_chart(chart, f'D{start_row}')

def format_answer_for_excel(answer, question):
    """Форматирует ответ для Excel"""
    if not answer.value:
        return 'Нет ответа'
    
    if question.type in ['grid', 'checkbox_grid']:
        try:
            if answer.value.startswith('['):
                selected_options = json.loads(answer.value)
                return '; '.join([f"{opt.split('|')[0]} → {opt.split('|')[1]}" if '|' in opt else opt for opt in selected_options])
            elif '|' in answer.value:
                row, col = answer.value.split('|', 1)
                return f"{row} → {col}"
        except:
            pass
    elif question.type in ['multiple_choice', 'checkbox']:
        try:
            if answer.value.startswith('['):
                selected_options = json.loads(answer.value)
                return '; '.join(selected_options)
        except:
            pass
    
    return answer.value[:100] + ('...' if len(answer.value) > 100 else '')