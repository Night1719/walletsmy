from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
import os
from datetime import datetime
import json
from functools import wraps

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-here')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///surveys.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
migrate = Migrate(app, db)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Модели базы данных
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    can_create_surveys = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    surveys = db.relationship('Survey', backref='creator', lazy=True)

class Survey(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    is_anonymous = db.Column(db.Boolean, default=False)
    require_auth = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    creator_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    questions = db.relationship('Question', backref='survey', lazy=True, cascade='all, delete-orphan')
    responses = db.relationship('SurveyResponse', backref='survey', lazy=True, cascade='all, delete-orphan')

class Question(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    text = db.Column(db.Text, nullable=False)
    type = db.Column(db.String(20), nullable=False)  # text, multiple_choice, rating
    options = db.Column(db.Text)  # JSON для вариантов ответов
    survey_id = db.Column(db.Integer, db.ForeignKey('survey.id'), nullable=False)
    
    answers = db.relationship('Answer', backref='question', lazy=True, cascade='all, delete-orphan')

class SurveyResponse(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    survey_id = db.Column(db.Integer, db.ForeignKey('survey.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    ip_address = db.Column(db.String(45), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    answers = db.relationship('Answer', backref='response', lazy=True, cascade='all, delete-orphan')

class Answer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    question_id = db.Column(db.Integer, db.ForeignKey('question.id'), nullable=False)
    response_id = db.Column(db.Integer, db.ForeignKey('survey_response.id'), nullable=False)
    value = db.Column(db.Text, nullable=False)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Вспомогательные функции для шаблонов
@app.context_processor
def utility_processor():
    def count_responses(surveys):
        """Подсчитывает общее количество ответов для списка опросов"""
        total = 0
        for survey in surveys:
            if hasattr(survey, 'responses'):
                total += len(survey.responses)
        return total
    
    def count_active_surveys(surveys):
        """Подсчитывает количество активных опросов (с ответами)"""
        count = 0
        for survey in surveys:
            if hasattr(survey, 'responses') and len(survey.responses) > 0:
                count += 1
        return count
    
    def format_date(date_obj):
        """Форматирует дату в удобочитаемый вид"""
        if date_obj:
            try:
                return date_obj.strftime('%d.%m')
            except:
                return str(date_obj)
        return '-'
    
    def from_json(json_string):
        """Преобразует JSON строку в Python объект"""
        try:
            if json_string:
                return json.loads(json_string)
            return []
        except (json.JSONDecodeError, TypeError):
            return []
    
    return {
        'count_responses': count_responses,
        'count_active_surveys': count_active_surveys,
        'format_date': format_date,
        'from_json': from_json
    }

# Регистрируем фильтр from_json отдельно
@app.template_filter('from_json')
def from_json_filter(json_string):
    """Фильтр для преобразования JSON строки в Python объект"""
    try:
        if json_string:
            return json.loads(json_string)
        return []
    except (json.JSONDecodeError, TypeError):
        return []

@app.template_filter('strftime')
def strftime_filter(date, format='%d.%m.%Y %H:%M'):
    """Фильтр для форматирования даты"""
    try:
        if date:
            return date.strftime(format)
        return ''
    except (AttributeError, TypeError):
        return ''

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('Требуются права администратора', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def survey_creation_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.can_create_surveys:
            flash('У вас нет прав на создание опросов', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# Маршруты
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            flash('Успешный вход!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Неверное имя пользователя или пароль', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Вы вышли из системы', 'info')
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.is_admin:
        surveys = Survey.query.all()
    else:
        surveys = Survey.query.filter_by(creator_id=current_user.id).all()
    
    return render_template('dashboard.html', surveys=surveys)

@app.route('/admin')
@admin_required
def admin_panel():
    users = User.query.all()
    surveys = Survey.query.all()
    
    # Используем SSL менеджер для получения реального статуса
    try:
        from ssl_manager import get_ssl_status
        ssl_status = get_ssl_status()
    except ImportError:
        # Fallback если SSL менеджер недоступен
        ssl_status = {
            'enabled': False,
            'certificate': None,
            'error': 'SSL менеджер недоступен'
        }
    
    return render_template('admin.html', users=users, surveys=surveys, ssl_status=ssl_status)

@app.route('/admin/users', methods=['GET', 'POST'])
@admin_required
def admin_users():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        is_admin = 'is_admin' in request.form
        can_create_surveys = 'can_create_surveys' in request.form
        
        if User.query.filter_by(username=username).first():
            flash('Пользователь с таким именем уже существует', 'error')
        else:
            user = User(
                username=username,
                email=email,
                password_hash=generate_password_hash(password),
                is_admin=is_admin,
                can_create_surveys=can_create_surveys
            )
            db.session.add(user)
            db.session.commit()
            flash('Пользователь создан успешно', 'success')
    
    users = User.query.all()
    return render_template('admin_users.html', users=users)

@app.route('/admin/users/<int:user_id>/toggle_admin')
@admin_required
def toggle_admin(user_id):
    user = User.query.get_or_404(user_id)
    if user.id != current_user.id:
        user.is_admin = not user.is_admin
        db.session.commit()
        flash(f'Права администратора для {user.username} изменены', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/toggle_survey_creation')
@admin_required
def toggle_survey_creation(user_id):
    user = User.query.get_or_404(user_id)
    user.can_create_surveys = not user.can_create_surveys
    db.session.commit()
    flash(f'Права на создание опросов для {user.username} изменены', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/delete')
@admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # Нельзя удалить самого себя
    if user.id == current_user.id:
        flash('Нельзя удалить свой собственный аккаунт', 'error')
        return redirect(url_for('admin_users'))
    
    # Нельзя удалить последнего администратора
    if user.is_admin:
        admin_count = User.query.filter_by(is_admin=True).count()
        if admin_count <= 1:
            flash('Нельзя удалить последнего администратора', 'error')
            return redirect(url_for('admin_users'))
    
    username = user.username
    
    # Удаляем пользователя (каскадное удаление)
    db.session.delete(user)
    db.session.commit()
    
    flash(f'Пользователь {username} удален успешно', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/ssl/upload', methods=['POST'])
@admin_required
def upload_ssl_certificate():
    """Загрузка SSL сертификата"""
    try:
        if 'certificate' not in request.files or 'private_key' not in request.files:
            return jsonify({'success': False, 'message': 'Необходимо загрузить оба файла'})
        
        cert_file = request.files['certificate']
        key_file = request.files['private_key']
        
        if cert_file.filename == '' or key_file.filename == '':
            return jsonify({'success': False, 'message': 'Файлы не выбраны'})
        
        # Проверяем расширения
        if not cert_file.filename.endswith('.pem') and not cert_file.filename.endswith('.crt'):
            return jsonify({'success': False, 'message': 'Сертификат должен быть в формате .pem или .crt'})
        
        if not key_file.filename.endswith('.pem') and not key_file.filename.endswith('.key'):
            return jsonify({'success': False, 'message': 'Приватный ключ должен быть в формате .pem или .key'})
        
        # Создаем папку ssl если её нет
        ssl_dir = 'ssl'
        if not os.path.exists(ssl_dir):
            os.makedirs(ssl_dir)
            print(f"✅ Создана папка {ssl_dir}")
        
        # Сохраняем файлы
        cert_path = os.path.join(ssl_dir, 'cert.pem')
        key_path = os.path.join(ssl_dir, 'key.pem')
        
        print(f"💾 Сохранение сертификата в: {cert_path}")
        print(f"💾 Сохранение ключа в: {key_path}")
        
        cert_file.save(cert_path)
        key_file.save(key_path)
        
        # Устанавливаем правильные права доступа
        os.chmod(key_path, 0o600)
        os.chmod(cert_path, 0o644)
        
        print(f"✅ Файлы сохранены с правами:")
        print(f"   Сертификат: {oct(os.stat(cert_path).st_mode)[-3:]}")
        print(f"   Ключ: {oct(os.stat(key_path).st_mode)[-3:]}")
        
        flash('SSL сертификат успешно загружен! Перезапустите сервер для применения изменений.', 'success')
        return jsonify({'success': True, 'message': 'SSL сертификат загружен успешно'})
        
    except Exception as e:
        print(f"❌ Ошибка загрузки SSL: {e}")
        return jsonify({'success': False, 'message': f'Ошибка загрузки: {str(e)}'})

@app.route('/admin/ssl/generate', methods=['POST'])
@admin_required
def generate_self_signed_certificate():
    """Генерация самоподписанного сертификата"""
    try:
        from ssl_manager import SSLManager
        
        ssl_manager = SSLManager()
        result = ssl_manager.generate_self_signed()
        
        if result['success']:
            flash('Самоподписанный сертификат успешно создан! Перезапустите сервер для применения изменений.', 'success')
            return jsonify({'success': True, 'message': 'Сертификат создан успешно'})
        else:
            return jsonify({'success': False, 'message': result['message']})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Ошибка генерации: {str(e)}'})

@app.route('/admin/ssl/letsencrypt', methods=['POST'])
@admin_required
def setup_lets_encrypt():
    """Настройка Let's Encrypt сертификата"""
    try:
        data = request.get_json()
        domain = data.get('domain')
        email = data.get('email')
        
        if not domain or not email:
            return jsonify({'success': False, 'message': 'Необходимо указать домен и email'})
        
        from ssl_manager import SSLManager
        
        ssl_manager = SSLManager()
        result = ssl_manager.setup_lets_encrypt(domain, email)
        
        if result['success']:
            flash('Let\'s Encrypt сертификат успешно настроен! Перезапустите сервер для применения изменений.', 'success')
            return jsonify({'success': True, 'message': 'Сертификат настроен успешно'})
        else:
            return jsonify({'success': False, 'message': result['message']})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Ошибка настройки: {str(e)}'})

@app.route('/admin/ssl/text-upload', methods=['POST'])
@admin_required
def upload_ssl_text():
    """Загрузка SSL сертификата текстом"""
    try:
        data = request.get_json()
        certificate = data.get('certificate', '').strip()
        private_key = data.get('private_key', '').strip()
        
        if not certificate or not private_key:
            return jsonify({'success': False, 'message': 'Необходимо заполнить оба поля'})
        
        # Проверяем формат сертификата
        if not certificate.startswith('-----BEGIN CERTIFICATE-----') or not certificate.endswith('-----END CERTIFICATE-----'):
            return jsonify({'success': False, 'message': 'Неверный формат сертификата'})
        
        # Проверяем формат ключа
        if not (private_key.startswith('-----BEGIN PRIVATE KEY-----') or 
                private_key.startswith('-----BEGIN RSA PRIVATE KEY-----')) or not private_key.endswith('-----END PRIVATE KEY-----'):
            return jsonify({'success': False, 'message': 'Неверный формат приватного ключа'})
        
        # Используем новый SSL менеджер
        from simple_ssl import ssl_manager
        
        if ssl_manager.save_certificate(certificate, private_key):
            return jsonify({'success': True, 'message': 'SSL сертификат успешно сохранен! Перезапустите сервер для применения изменений.'})
        else:
            return jsonify({'success': False, 'message': 'Ошибка сохранения SSL файлов'})
        
    except Exception as e:
        print(f"❌ Ошибка сохранения SSL: {e}")
        return jsonify({'success': False, 'message': f'Ошибка сохранения: {str(e)}'})

@app.route('/surveys/create', methods=['GET', 'POST'])
@login_required
@survey_creation_required
def create_survey():
    if request.method == 'POST':
        title = request.form.get('title')
        description = request.form.get('description')
        is_anonymous = 'is_anonymous' in request.form
        require_auth = 'require_auth' in request.form
        
        survey = Survey(
            title=title,
            description=description,
            is_anonymous=is_anonymous,
            require_auth=require_auth,
            creator_id=current_user.id
        )
        db.session.add(survey)
        db.session.commit()
        
        # Добавляем вопросы
        questions_data = json.loads(request.form.get('questions', '[]'))
        for q_data in questions_data:
            question = Question(
                text=q_data['text'],
                type=q_data['type'],
                options=json.dumps(q_data.get('options', [])),
                survey_id=survey.id
            )
            db.session.add(question)
        
        db.session.commit()
        flash('Опрос создан успешно', 'success')
        return redirect(url_for('dashboard'))
    
    return render_template('create_survey.html')

@app.route('/surveys/<int:survey_id>')
def view_survey(survey_id):
    survey = Survey.query.get_or_404(survey_id)
    return render_template('view_survey.html', survey=survey)

@app.route('/surveys/<int:survey_id>/submit', methods=['POST'])
def submit_survey(survey_id):
    survey = Survey.query.get_or_404(survey_id)
    
    if survey.require_auth and not current_user.is_authenticated:
        flash('Для прохождения этого опроса требуется авторизация', 'error')
        return redirect(url_for('login'))
    
    # Создаем ответ на опрос
    response = SurveyResponse(
        survey_id=survey.id,
        user_id=current_user.id if current_user.is_authenticated else None,
        ip_address=request.remote_addr
    )
    db.session.add(response)
    db.session.commit()
    
    # Сохраняем ответы на вопросы
    for question in survey.questions:
        answer_value = request.form.get(f'question_{question.id}')
        if answer_value:
            answer = Answer(
                question_id=question.id,
                response_id=response.id,
                value=answer_value
            )
            db.session.add(answer)
    
    db.session.commit()
    flash('Опрос успешно пройден!', 'success')
    return redirect(url_for('index'))

@app.route('/surveys/<int:survey_id>/results')
@login_required
def survey_results(survey_id):
    survey = Survey.query.get_or_404(survey_id)
    
    if not current_user.is_admin and survey.creator_id != current_user.id:
        flash('У вас нет доступа к результатам этого опроса', 'error')
        return redirect(url_for('dashboard'))
    
    # Получаем все ответы на опрос
    responses = SurveyResponse.query.filter_by(survey_id=survey_id).order_by(SurveyResponse.created_at.desc()).all()
    
    # Анализ общих результатов
    results = {}
    for question in survey.questions:
        try:
            if question.type == 'multiple_choice':
                # Безопасно загружаем опции
                try:
                    if question.options and question.options.strip():
                        options = json.loads(question.options)
                        if isinstance(options, list) and options:
                            counts = {opt: 0 for opt in options}
                            for answer in question.answers:
                                if answer.value in counts:
                                    counts[answer.value] += 1
                            results[question.id] = {
                                'type': 'multiple_choice',
                                'text': question.text,
                                'data': counts
                            }
                        else:
                            # Если опции пустые или невалидные
                            results[question.id] = {
                                'type': 'multiple_choice',
                                'text': question.text,
                                'data': {},
                                'error': 'Варианты ответов не настроены'
                            }
                    else:
                        # Если опции отсутствуют
                        results[question.id] = {
                            'type': 'multiple_choice',
                            'text': question.text,
                            'data': {},
                            'error': 'Варианты ответов не настроены'
                        }
                except (json.JSONDecodeError, TypeError) as e:
                    print(f"❌ Ошибка парсинга опций для вопроса {question.id}: {e}")
                    results[question.id] = {
                        'type': 'multiple_choice',
                        'text': question.text,
                        'data': {},
                        'error': 'Ошибка в настройке вариантов ответов'
                    }
                    
            elif question.type == 'rating':
                ratings = []
                for answer in question.answers:
                    try:
                        if answer.value and answer.value.isdigit():
                            ratings.append(int(answer.value))
                    except (ValueError, TypeError):
                        continue
                
                if ratings:
                    # Создаем данные для графика рейтингов
                    rating_data = {}
                    for rating in range(1, 11):
                        rating_data[rating] = ratings.count(rating)
                    
                    results[question.id] = {
                        'type': 'rating',
                        'text': question.text,
                        'average': sum(ratings) / len(ratings),
                        'count': len(ratings),
                        'data': rating_data
                    }
                else:
                    results[question.id] = {
                        'type': 'rating',
                        'text': question.text,
                        'average': 0,
                        'count': 0,
                        'data': {},
                        'error': 'Нет валидных ответов'
                    }
                    
            else:  # text
                answers = []
                for answer in question.answers:
                    if answer.value and answer.value.strip():
                        answers.append(answer.value.strip())
                
                results[question.id] = {
                    'type': 'text',
                    'text': question.text,
                    'answers': answers
                }
                
        except Exception as e:
            print(f"❌ Ошибка обработки вопроса {question.id}: {e}")
            results[question.id] = {
                'type': 'error',
                'text': question.text,
                'error': f'Ошибка обработки: {str(e)}'
            }
    
    return render_template('survey_results.html', survey=survey, results=results, responses=responses)

@app.route('/surveys/<int:survey_id>/export-excel')
@login_required
def export_survey_excel(survey_id):
    """Экспорт результатов опроса в Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from flask import send_file
        
        survey = Survey.query.get_or_404(survey_id)
        
        if not current_user.is_admin and survey.creator_id != current_user.id:
            flash('У вас нет доступа к результатам этого опроса', 'error')
            return redirect(url_for('dashboard'))
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты опроса"
        
        # Заголовки
        headers = ['№', 'Вопрос', 'Тип', 'Ответы', 'Статистика']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Данные
        row = 2
        for question in survey.questions:
            # Номер вопроса
            ws.cell(row=row, column=1, value=row-1)
            
            # Текст вопроса
            ws.cell(row=row, column=2, value=question.text)
            
            # Тип вопроса
            type_names = {
                'text': 'Текстовый',
                'multiple_choice': 'Множественный выбор',
                'rating': 'Рейтинг'
            }
            ws.cell(row=row, column=3, value=type_names.get(question.type, question.type))
            
            # Ответы
            if question.type == 'multiple_choice':
                try:
                    options = json.loads(question.options) if question.options else []
                    answers_text = ', '.join(options) if options else 'Варианты не настроены'
                except:
                    answers_text = 'Ошибка в вариантах'
            elif question.type == 'rating':
                answers_text = 'Рейтинг от 1 до 10'
            else:
                answers_text = 'Текстовые ответы'
            
            ws.cell(row=row, column=4, value=answers_text)
            
            # Статистика
            if question.type == 'multiple_choice':
                try:
                    options = json.loads(question.options) if question.options else []
                    if options:
                        counts = {opt: 0 for opt in options}
                        for answer in question.answers:
                            if answer.value in counts:
                                counts[answer.value] += 1
                        stats = '; '.join([f"{opt}: {count}" for opt, count in counts.items()])
                    else:
                        stats = 'Нет вариантов'
                except:
                    stats = 'Ошибка в данных'
            elif question.type == 'rating':
                ratings = [int(answer.value) for answer in question.answers if answer.value and answer.value.isdigit()]
                if ratings:
                    avg = sum(ratings) / len(ratings)
                    stats = f"Средний: {avg:.1f}, Всего: {len(ratings)}"
                else:
                    stats = 'Нет ответов'
            else:
                answers_count = len([a for a in question.answers if a.value])
                stats = f"Ответов: {answers_count}"
            
            ws.cell(row=row, column=5, value=stats)
            
            row += 1
        
        # Автоматическая ширина столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f"survey_results_{survey.title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"❌ Ошибка экспорта в Excel: {e}")
        flash('Ошибка экспорта в Excel', 'error')
        return redirect(url_for('survey_results', survey_id=survey_id))

@app.route('/surveys/<int:survey_id>/response/<int:response_id>')
@login_required
def view_response_detail(survey_id, response_id):
    """Просмотр детального ответа пользователя"""
    survey = Survey.query.get_or_404(survey_id)
    response = SurveyResponse.query.get_or_404(response_id)
    
    if not current_user.is_admin and survey.creator_id != current_user.id:
        flash('У вас нет доступа к результатам этого опроса', 'error')
        return redirect(url_for('dashboard'))
    
    if response.survey_id != survey_id:
        flash('Ответ не принадлежит указанному опросу', 'error')
        return redirect(url_for('survey_results', survey_id=survey_id))
    
    # Получаем ответы на вопросы
    answers = {}
    for answer in response.answers:
        question = answer.question
        if question:
            answers[question.id] = {
                'question_text': question.text,
                'question_type': question.type,
                'answer_value': answer.value,
                'question_options': question.options
            }
    
    return render_template('response_detail.html', 
                         survey=survey, 
                         response=response, 
                         answers=answers)

# LDAP маршруты
@app.route('/admin/ldap/test')
@admin_required
def test_ldap_connection():
    """Тестирование подключения к LDAP"""
    try:
        from ldap_manager import ldap_manager
        result = ldap_manager.test_connection()
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': f'Ошибка: {str(e)}'})

@app.route('/admin/ldap/search')
@admin_required
def search_ldap_users():
    """Поиск пользователей в LDAP"""
    try:
        from ldap_manager import ldap_manager
        query = request.args.get('q', '')
        max_results = int(request.args.get('max', 50))
        
        users = ldap_manager.search_users(query, max_results)
        return jsonify({'success': True, 'users': users})
    except Exception as e:
        return jsonify({'success': False, 'error': f'Ошибка: {str(e)}'})

@app.route('/admin/ldap/import', methods=['POST'])
@admin_required
def import_ldap_users():
    """Импорт пользователей из LDAP"""
    try:
        from ldap_manager import ldap_manager
        from werkzeug.security import generate_password_hash
        
        data = request.get_json()
        user_dns = data.get('user_dns', [])
        
        if not user_dns:
            return jsonify({'success': False, 'error': 'Не выбраны пользователи для импорта'})
        
        # Импортируем пользователей
        result = ldap_manager.import_users(user_dns)
        
        if result['success']:
            # Создаем пользователей в системе
            created_count = 0
            for user_data in result['imported_users']:
                # Проверяем, не существует ли уже пользователь
                existing_user = User.query.filter_by(username=user_data['username']).first()
                if not existing_user:
                    # Создаем нового пользователя
                    new_user = User(
                        username=user_data['username'],
                        email=user_data['email'] or f"{user_data['username']}@buntergroup.com",
                        password_hash=generate_password_hash('changeme123'),  # Временный пароль
                        is_admin=False,
                        can_create_surveys=False
                    )
                    db.session.add(new_user)
                    created_count += 1
            
            if created_count > 0:
                db.session.commit()
                flash(f'Импортировано {created_count} пользователей из LDAP', 'success')
            else:
                flash('Все выбранные пользователи уже существуют в системе', 'info')
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Ошибка импорта: {str(e)}'})

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    
    # Проверяем SSL и запускаем соответственно
    try:
        from simple_ssl import ssl_manager
        
        print("🔍 Проверка SSL конфигурации...")
        
        if ssl_manager.is_ssl_ready():
            print("✅ SSL статус: Включен")
            ssl_context = ssl_manager.get_ssl_context()
            
            if ssl_context:
                print("🔒 Запуск с SSL сертификатом...")
                print(f"   Порт: 5000 (HTTPS)")
                print(f"   Сертификат: {ssl_manager.cert_file}")
                print(f"   Ключ: {ssl_manager.key_file}")
                
                app.run(debug=True, host='0.0.0.0', port=5000, ssl_context=ssl_context)
            else:
                print("⚠️  SSL файлы найдены, но не валидны. Запуск без SSL...")
                print("   Проверьте права доступа и формат файлов")
                app.run(debug=True, host='0.0.0.0', port=5000)
        else:
            print("🌐 Запуск без SSL...")
            print("   Проверьте наличие файлов cert.pem и key.pem в папке ssl/")
            app.run(debug=True, host='0.0.0.0', port=5000)
            
    except ImportError as e:
        print(f"⚠️  SSL менеджер недоступен: {e}")
        print("   Запуск без SSL...")
        app.run(debug=True, host='0.0.0.0', port=5000)
    except Exception as e:
        print(f"❌ Ошибка SSL конфигурации: {e}")
        print("   Запуск без SSL...")
        app.run(debug=True, host='0.0.0.0', port=5000)